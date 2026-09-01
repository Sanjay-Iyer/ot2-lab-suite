"""exp4_scan_key.py - the bench-readable scan key for the 083126 dump.

One row per scan: what was on the paper, which dye dilution, how long the
exposure was, and whether the saturation screen let it into the gains. Written
as a flat CSV and as a four-tab workbook split by particle and paper.

Generated from `configs/stars_exp4.yaml` and the screen table rather than kept
by hand, so a corrected `dilutions:` list - as on 090126, when the bipyramid
rerun turned out to be 2x/5x/10x and not 2x/10x/20x - reaches the key and the
figures in the same run. Called by exp4_stars_report.main().
"""
from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl.utils import get_column_letter

PAPER_LABEL = {"offwhite": "Off-white", "white": "White", "?": "Unknown"}
PARTICLE_LABEL = {"stars": "Nanostars", "bipyramids": "Bipyramids", "?": "Unassigned"}
SAMPLE_LABEL = {
    "stars": "Stock nanostars + CV",
    "stars5x": "5:1 diluted nanostars + CV",
    "bp": "Bipyramids + CV",
    "dyeonly": "CV only",
}
PANEL_LABEL = {
    "ow-10": "Off-white nanostars, 10 s",
    "w-10": "White nanostars, 10 s",
    "w-6": "White nanostars, 6 s",
    "bp-w-10": "White bipyramids, 10 s",
    "bp-w-6": "White bipyramids, 6 s",
    "bp-ow-10": "Off-white bipyramids, 10 s",
    "unassigned": "Unassigned",
}
COLUMNS = [
    "scan_number", "paper", "nanoparticle", "sample", "cv_dilution",
    "exposure_s", "panel", "panel_key", "quality_verdict",
    "quality_verdict_whole_scan", "usable_for_gain", "raw_folder", "acquired_at",
    "notes",
]
TABS = (
    ("Stars - Off-white", "Nanostars", "Off-white"),
    ("Stars - White", "Nanostars", "White"),
    ("Bipyramids - Off-white", "Bipyramids", "Off-white"),
    ("Bipyramids - White", "Bipyramids", "White"),
)


def _nanoparticle(condition, particle):
    """What was actually on the spot.

    A dye-only control has no nanoparticles on it - it is dye on bare paper -
    so it is not labelled with the particle of the block it belongs to. The
    block is still recoverable from `panel` / `panel_key`, which is what the
    tabs group on.
    """
    if condition == "dyeonly":
        return "None (dye only)"
    return PARTICLE_LABEL.get(particle, particle)


def _block_particle(panel_key):
    """The particle a panel is a panel OF, dye-only controls included."""
    return "Bipyramids" if str(panel_key).startswith("bp-") else "Nanostars"


def _missing_rows(cfg, screen) -> list[dict]:
    """Placeholder rows for scans the config maps but the dump does not hold.

    They are carried rather than dropped so the key shows the hole where the
    white 10 s strong end should be, instead of a series that silently starts
    at 50x.
    """
    present = set(screen["scan_number"])
    rows = []
    for scan, m in (cfg.get("missing_scans") or {}).items():
        if int(scan) in present:
            continue
        rows.append({
            "scan_number": int(scan),
            "paper": PAPER_LABEL.get(m["paper"], m["paper"]),
            "nanoparticle": _nanoparticle(m["condition"], m["particle"]),
            "sample": SAMPLE_LABEL.get(m["condition"], m["condition"]),
            "cv_dilution": m["dilution"],
            "exposure_s": pd.NA,
            "panel": PANEL_LABEL.get(m["key"], m["key"]),
            "panel_key": m["key"],
            "quality_verdict": "missing raw file",
            "quality_verdict_whole_scan": "missing raw file",
            "usable_for_gain": "no",
            "raw_folder": m["folder"],
            "acquired_at": pd.NA,
            "notes": m.get("note", "Expected scan in this series; the raw file "
                            "is absent from the current data dump."),
        })
    return rows


def build(cfg: dict, screen: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for _, r in screen.iterrows():
        condition = r["condition"] if isinstance(r["condition"], str) else None
        rows.append({
            "scan_number": int(r["scan_number"]),
            "paper": PAPER_LABEL.get(r["paper"], r["paper"]),
            "nanoparticle": (_nanoparticle(condition, r["particle"])
                             if condition else PARTICLE_LABEL.get(r["particle"],
                                                                  r["particle"])),
            "sample": SAMPLE_LABEL.get(condition, "Unassigned probe"),
            "cv_dilution": r["dilution"] if isinstance(r["dilution"], str) else pd.NA,
            "exposure_s": float(r["int_time_s"]),
            "panel": PANEL_LABEL.get(r["key"], r["key"]),
            "panel_key": r["key"],
            "quality_verdict": r["verdict"],
            "quality_verdict_whole_scan": r["verdict_scan"],
            # A failed scan is excluded from every gain; warn and pass are not.
            "usable_for_gain": "no" if (r["verdict"] == "fail" or condition is None)
                               else "yes",
            "raw_folder": r["folder"],
            "acquired_at": r["created"],
            "notes": r["series_note"] or "",
        })
    rows.extend(_missing_rows(cfg, screen))
    key = pd.DataFrame(rows, columns=COLUMNS)
    return key.sort_values("scan_number").reset_index(drop=True)


def _write_tab(writer, sheet: str, subtitle: str, frame: pd.DataFrame) -> None:
    """Title, one-line subtitle, blank row, then the table."""
    header = pd.DataFrame([[sheet] + [None] * (len(COLUMNS) - 1),
                           [subtitle] + [None] * (len(COLUMNS) - 1),
                           [None] * len(COLUMNS)])
    header.to_excel(writer, sheet_name=sheet, index=False, header=False)
    frame.to_excel(writer, sheet_name=sheet, index=False, startrow=3)
    sheet_obj = writer.sheets[sheet]
    for i, column in enumerate(COLUMNS):
        width = max(len(column), int(frame[column].astype(str).str.len().max() or 0))
        sheet_obj.column_dimensions[get_column_letter(i + 1)].width = (
            min(max(width + 2, 10), 60)
        )


def _fallback(path: Path) -> Path:
    """Where a locked output goes instead: name.updated.ext, same folder."""
    return path.with_suffix(".updated" + path.suffix)


def _write_csv(key: pd.DataFrame, path: Path) -> Path:
    try:
        key.to_csv(path, index=False)
        return path
    except PermissionError:
        alt = _fallback(path)
        key.to_csv(alt, index=False)
        return alt


def _write_workbook(key: pd.DataFrame, path: Path) -> Path:
    def render(target: Path) -> None:
        with pd.ExcelWriter(target, engine="openpyxl") as writer:
            for sheet, particle, paper in TABS:
                frame = key[
                    (key["panel_key"].map(_block_particle) == particle)
                    & (key["paper"] == paper)
                ].reset_index(drop=True)
                _write_tab(
                    writer, sheet,
                    "%s on %s paper - %d scans, CV dilutions %s."
                    % (particle, paper.lower(), len(frame),
                       ", ".join(dict.fromkeys(frame["cv_dilution"].dropna()))),
                    frame,
                )

    try:
        render(path)
        return path
    except PermissionError:
        alt = _fallback(path)
        render(alt)
        return alt


def write(cfg: dict, screen: pd.DataFrame, out_dir: Path) -> pd.DataFrame:
    """Write the key, routing around a file open in Excel rather than failing.

    A locked output is written next to the original as `name.updated.ext` and
    named in the console, so a corrected run is never silently lost to a file
    lock - and rerunning once Excel is closed puts it back in the real place.
    """
    key = build(cfg, screen)
    written = [
        _write_csv(key, out_dir / "scan_key.csv"),
        _write_workbook(key, out_dir / "scan_key_4_tabs.xlsx"),
    ]
    diverted = [p for p in written if p.name.endswith((".updated.csv", ".updated.xlsx"))]
    if diverted:
        print("[exp4] NOTE: %d scan-key file(s) were open in Excel and could not "
              "be replaced. The corrected copies are %s - close Excel and rerun "
              "to write them in place."
              % (len(diverted), ", ".join(p.name for p in diverted)))
    return key


# ---------------------------------------------------------------------------
# v2: the plotted subset
#
# v1 is the whole dump - every scan, every rung, placeholders for the files
# that never arrived. v2 answers a narrower question: which scans are behind
# the figures? That is one block per particle per paper, held to the dilutions
# every arm shares, with the quantity the figures actually draw (the 1620 cm-1
# peak height) and the gain each spot earns over the dye-only control beside
# it. Saturated scans stay in, flagged: they are drawn on the bar figure and
# dropped from the spectra, and a row that vanished silently would be worse
# than a row that says why it is there.
# ---------------------------------------------------------------------------
V2_COLUMNS = [
    "scan_number", "paper", "nanoparticle", "sample", "cv_dilution",
    "exposure_s", "panel", "verdict_at_band", "verdict_whole_scan",
    "noise_excess_at_band", "counts_at_band", "in_plots",
    "height_1620_cps", "snr_1620", "gain_vs_dye_x", "raw_folder", "notes",
]
CONDITION_ORDER = {"stars": 0, "stars5x": 1, "bp": 2, "dyeonly": 3}


def _active_blocks(cfg) -> list[tuple[str, str, str]]:
    """(sheet, panel key, paper) for the four blocks the figures are built on."""
    blocks = []
    for paper, sheet in (("offwhite", "Off-white"), ("white", "White")):
        blocks.append(("Stars - %s" % sheet,
                       cfg["plotting"]["paper_focus"][paper]["key"], paper))
        blocks.append(("Bipyramids - %s" % sheet,
                       cfg["plotting"]["particle_comparison"][paper]["bipyramids_key"],
                       paper))
    return blocks


def build_v2(cfg, screen, bands, gain) -> dict[str, pd.DataFrame]:
    active = cfg["plotting"].get("active_dilutions")
    band = cfg["sers_band"]
    metrics = (bands[bands["peak_name"] == band]
               .set_index("scan_number")[["height_cps", "snr"]])
    gains = {}
    if len(gain):
        for _, g in gain[gain["peak_name"] == band].iterrows():
            gains[(g["key"], g["condition"], g["dilution"])] = g["gain_x"]

    sheets = {}
    for sheet, key, paper in _active_blocks(cfg):
        rows = screen[(screen["key"] == key) & screen["condition"].notna()]
        if active:
            rows = rows[rows["dilution"].isin(active)]
        controls = {
            r["dilution"]: r for _, r in rows.iterrows()
            if r["condition"] == "dyeonly"
        }
        out = []
        for _, r in rows.iterrows():
            scan = int(r["scan_number"])
            m = metrics.loc[scan] if scan in metrics.index else None
            out.append({
                "scan_number": scan,
                "paper": PAPER_LABEL.get(r["paper"], r["paper"]),
                "nanoparticle": _nanoparticle(r["condition"], r["particle"]),
                "sample": SAMPLE_LABEL.get(r["condition"], r["condition"]),
                "cv_dilution": r["dilution"],
                "exposure_s": float(r["int_time_s"]),
                "panel": PANEL_LABEL.get(r["key"], r["key"]),
                "verdict_at_band": r["verdict_band"],
                "verdict_whole_scan": r["verdict_scan"],
                "noise_excess_at_band": round(float(r["noise_excess_at_band"]), 2),
                "counts_at_band": round(float(r["level_at_band_counts"])),
                "in_plots": _in_plots(r),
                "height_1620_cps": None if m is None else round(float(m["height_cps"]), 2),
                "snr_1620": None if m is None else round(float(m["snr"]), 1),
                # The control is the thing gains are measured against, so it has
                # no gain of its own; a saturated spot is excluded from gains.
                "gain_vs_dye_x": (None if r["condition"] == "dyeonly" else
                                  _round(gains.get((key, r["condition"], r["dilution"])))),
                "raw_folder": r["folder"],
                "notes": _note(r, gains, controls, key),
            })
        frame = pd.DataFrame(out, columns=V2_COLUMNS)
        frame["_c"] = frame["sample"].map(
            {v: CONDITION_ORDER[k] for k, v in SAMPLE_LABEL.items()}).fillna(9)
        frame["_d"] = frame["cv_dilution"].map(lambda d: float(str(d).rstrip("x")))
        sheets[sheet] = (frame.sort_values(["_c", "_d"])
                         .drop(columns=["_c", "_d"]).reset_index(drop=True))
    return sheets


def _in_plots(row):
    """Whether this scan reaches the figures, and on what terms.

    Three states, and an empty cell would flatten all of them: out because the
    band itself is noise; in, but ringing somewhere the band is not; in, but
    sitting high enough on the detector that the height is a lower bound.
    """
    if row["verdict_band"] == "fail":
        return "no - the band itself is in the noise"
    parts = ["yes"]
    if row["verdict_scan"] == "fail":
        parts.append("rings below ~800 cm-1, clean at the band")
    if row["band_above_fit_ceiling"]:
        parts.append("counts above the linearity ceiling - height is a lower bound")
    return "; ".join(parts)


def _note(row, gains, controls, key):
    """The bench note, plus the reason a gain cell is blank when it is.

    A blank gain on a test spot is not the same fact as a blank gain on the
    control, and neither is visible from an empty cell, so the reason is
    written where the reader is already looking.
    """
    note = row["series_note"] or ""
    if row["condition"] == "dyeonly":
        return note
    if not pd.isna(gains.get((key, row["condition"], row["dilution"]), np.nan)):
        return note
    control = controls.get(row["dilution"])
    if control is None:
        reason = "no gain: this block has no dye-only control at %s" % row["dilution"]
    elif control["verdict"] == "fail" and row["verdict"] == "fail":
        reason = ("no gain: this scan and the %s dye-only control (scan %d) are "
                  "both saturated" % (row["dilution"], int(control["scan_number"])))
    elif control["verdict"] == "fail":
        reason = ("no gain: the %s dye-only control (scan %d) is saturated"
                  % (row["dilution"], int(control["scan_number"])))
    elif row["verdict"] == "fail":
        reason = "no gain: this scan is saturated and is excluded from gains"
    else:
        reason = "no gain recorded at this dilution"
    return "%s; %s" % (note, reason) if note else reason


def _round(value):
    return None if value is None or pd.isna(value) else round(float(value), 2)


def write_v2(cfg, screen, bands, gain, out_dir: Path) -> Path:
    sheets = build_v2(cfg, screen, bands, gain)
    path = out_dir / "scan_key_v2_plot_data_4_tabs.xlsx"

    def render(target: Path) -> None:
        with pd.ExcelWriter(target, engine="openpyxl") as writer:
            for sheet, frame in sheets.items():
                dilutions = ", ".join(dict.fromkeys(frame["cv_dilution"]))
                saturated = int((frame["verdict_at_band"] == "fail").sum())
                subtitle = (
                    "%s - the %d scans behind this block's figures, at %s. "
                    "Gain is each spot over the CV-only control at the same "
                    "dilution in the same block.%s"
                    % (frame["panel"].iloc[0], len(frame), dilutions,
                       "" if not saturated else
                       " %d saturated scan(s) are kept and flagged." % saturated)
                )
                _write_tab_v2(writer, sheet, subtitle, frame)

    try:
        render(path)
    except PermissionError:
        path = _fallback(path)
        render(path)
        print("[exp4] NOTE: the v2 workbook was open in Excel; wrote %s instead"
              % path.name)
    return path


def _write_tab_v2(writer, sheet: str, subtitle: str, frame: pd.DataFrame) -> None:
    header = pd.DataFrame([[sheet] + [None] * (len(V2_COLUMNS) - 1),
                           [subtitle] + [None] * (len(V2_COLUMNS) - 1),
                           [None] * len(V2_COLUMNS)])
    header.to_excel(writer, sheet_name=sheet, index=False, header=False)
    frame.to_excel(writer, sheet_name=sheet, index=False, startrow=3)
    sheet_obj = writer.sheets[sheet]
    for i, column in enumerate(V2_COLUMNS):
        width = max(len(column), int(frame[column].astype(str).str.len().max() or 0))
        sheet_obj.column_dimensions[get_column_letter(i + 1)].width = (
            min(max(width + 2, 10), 60)
        )
